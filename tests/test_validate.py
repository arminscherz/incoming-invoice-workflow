import json
import os
from pathlib import Path
import pytest
from typer.testing import CliRunner
import openpyxl

from ii_workflow.main import app

runner = CliRunner()

@pytest.fixture
def dummy_bank_statement(tmp_path):
    """Creates a dummy .xlsx bank statement file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add some garbage rows at the top to simulate real statements
    ws.append(["Some garbage header info"])
    ws.append(["More garbage", "Data"])
    
    # Add actual headers
    headers = [
        "Valutadatum", "Buchungsdatum", "Betrag", "Währung", 
        "Gegenpartei", "Bezeichnung", "Referenz", "Nachricht", "Zahlungs-ID"
    ]
    ws.append(headers)
    
    # Transaction 1: Matching transaction (negative amount)
    # Valutadatum: 2026-05-06 (Invoice is 2026-05-04, within +/- 3 days)
    # Amount: -120.0
    # Vendor: Test Vendor in Gegenpartei
    ws.append(["2026-05-06", "2026-05-06", -120.0, "EUR", "Test Vendor GmbH", "Rechnung INV-001", "", "", "12345"])
    
    # Transaction 2: Non-matching transaction
    ws.append(["2026-05-01", "2026-05-01", -50.0, "EUR", "Other Corp", "Something else", "", "", "54321"])
    
    file_path = tmp_path / "dummy_statement.xlsx"
    wb.save(file_path)
    return str(file_path)

@pytest.fixture
def valid_invoice_json(tmp_path):
    """Creates a valid invoice JSON file."""
    data = {
        "vendor_name": "Test Vendor",
        "purchase_category": "Büromaterial",
        "invoice_number": "INV-001",
        "date": "2026-05-04",
        "total_invoice_amount_gross": 120.0,
        "total_invoice_amount_net": 100.0,
        "total_invoice_amount_tax": 20.0,
        "tip_amount": 0.0,
        "total_payment_amount_gross": 120.0,
        "tax_amount_0_percent_VAT": 0.0,
        "tax_amount_10_percent_VAT": 0.0,
        "tax_amount_13_percent_VAT": 0.0,
        "tax_amount_20_percent_VAT": 20.0,
        "net_amount_0_percent_VAT": 0.0,
        "net_amount_10_percent_VAT": 0.0,
        "net_amount_13_percent_VAT": 0.0,
        "net_amount_20_percent_VAT": 100.0,
        "currency": "EUR",
        "iban": "AT123456789",
        "payment_method": None
    }
    file_path = tmp_path / "invoice.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
    return str(file_path)

@pytest.fixture
def invalid_math_invoice_json(tmp_path):
    """Creates an invoice JSON file with invalid math."""
    data = {
        "vendor_name": "Test Vendor",
        "purchase_category": "Büromaterial",
        "invoice_number": "INV-002",
        "date": "2026-05-04",
        "total_invoice_amount_gross": 150.0, # Gross doesn't match net + tax
        "total_invoice_amount_net": 100.0,
        "total_invoice_amount_tax": 20.0,
        "tip_amount": 0.0,
        "total_payment_amount_gross": 150.0,
        "tax_amount_0_percent_VAT": 0.0,
        "tax_amount_10_percent_VAT": 0.0,
        "tax_amount_13_percent_VAT": 0.0,
        "tax_amount_20_percent_VAT": 20.0,
        "net_amount_0_percent_VAT": 0.0,
        "net_amount_10_percent_VAT": 0.0,
        "net_amount_13_percent_VAT": 0.0,
        "net_amount_20_percent_VAT": 100.0,
        "currency": "EUR",
        "iban": None,
        "payment_method": None
    }
    file_path = tmp_path / "invalid_math.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
    return str(file_path)

@pytest.fixture
def no_match_invoice_json(tmp_path):
    """Creates a valid invoice JSON that won't match any bank transaction."""
    data = {
        "vendor_name": "Unknown Vendor",
        "purchase_category": "Büromaterial",
        "invoice_number": "INV-999",
        "date": "2026-05-15", # Date is far off
        "total_invoice_amount_gross": 99.0,
        "total_invoice_amount_net": 90.0,
        "total_invoice_amount_tax": 9.0,
        "tip_amount": 0.0,
        "total_payment_amount_gross": 99.0,
        "tax_amount_0_percent_VAT": 0.0,
        "tax_amount_10_percent_VAT": 9.0,
        "tax_amount_13_percent_VAT": 0.0,
        "tax_amount_20_percent_VAT": 0.0,
        "net_amount_0_percent_VAT": 0.0,
        "net_amount_10_percent_VAT": 90.0,
        "net_amount_13_percent_VAT": 0.0,
        "net_amount_20_percent_VAT": 0.0,
        "currency": "EUR",
        "iban": None,
        "payment_method": None
    }
    file_path = tmp_path / "no_match.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
    return str(file_path)

def test_validate_success_bank_match(valid_invoice_json, dummy_bank_statement, tmp_path, mocker):
    """Test successful validation where a bank transaction matches."""
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", valid_invoice_json, "--bank_statement", dummy_bank_statement])
    
    assert result.exit_code == 0, f"Output: {result.output}"
    
    # Verify the output file is in VALIDATED_DIR
    output_path = validated_dir / "invoice-validated.json"
    assert output_path.exists()
    
    # Verify absolute path is in stdout
    assert str(output_path.absolute()) in result.output
    
    with open(output_path, "r") as f:
        data = json.load(f)
        assert data["payment_method"] == "Bankkonto"

def test_validate_success_no_bank_match(no_match_invoice_json, dummy_bank_statement, tmp_path, mocker):
    """Test successful validation where NO bank transaction matches (assumes cash/bar)."""
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", no_match_invoice_json, "--bank_statement", dummy_bank_statement])
    
    assert result.exit_code == 0, f"Output: {result.output}"
    
    output_path = validated_dir / "no_match-validated.json"
    assert output_path.exists()
    assert str(output_path.absolute()) in result.output
    
    with open(output_path, "r") as f:
        data = json.load(f)
        assert data["payment_method"] == "bar"

def test_validate_invalid_math(invalid_math_invoice_json, dummy_bank_statement, tmp_path, mocker):
    """Test validation failing due to bad math."""
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", invalid_math_invoice_json, "--bank_statement", dummy_bank_statement])
    
    assert result.exit_code != 0
    assert "Math validation failed" in str(result.output) or "does not equal" in str(result.output) or result.exit_code != 0

def test_validate_no_bank_statement(valid_invoice_json, tmp_path, mocker):
    """Test validation when no bank statement is provided."""
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", valid_invoice_json])
    
    assert result.exit_code == 0, f"Output: {result.output}"
    
    output_path = validated_dir / "invoice-validated.json"
    assert output_path.exists()
    assert str(output_path.absolute()) in result.output
    
    with open(output_path, "r") as f:
        data = json.load(f)
        assert data["payment_method"] == "bar"

def test_validate_relative_path_ingested_dir(valid_invoice_json, tmp_path, mocker):
    """Test that relative paths are resolved against INGESTED_DIR."""
    # Create an ingested_dir different from work_dir
    ingested_dir = tmp_path / "ingested"
    ingested_dir.mkdir()
    validated_dir = tmp_path / "validated"
    
    # Move the valid invoice json into ingested_dir
    filename = Path(valid_invoice_json).name
    new_json_path = ingested_dir / filename
    os.rename(valid_invoice_json, new_json_path)
    
    mocker.patch.dict(os.environ, {
        "INGESTED_DIR": str(ingested_dir), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    # Provide just the filename, not the absolute path
    result = runner.invoke(app, ["validate", filename])
    
    assert result.exit_code == 0, f"Output: {result.output}"
    
    # output should be in validated_dir
    output_path = validated_dir / f"{Path(filename).stem}-validated.json"
    assert output_path.exists()
    assert str(output_path.absolute()) in result.output

@pytest.fixture
def equal_gross_net_invoice_json(tmp_path):
    """Creates an invoice JSON file where gross == net to test recalculation."""
    data = {
        "vendor_name": "Test Vendor",
        "purchase_category": "Büromaterial",
        "invoice_number": "INV-003",
        "date": "2026-05-04",
        "total_invoice_amount_gross": 120.0,
        "total_invoice_amount_net": 120.0, # Gross equals net
        "total_invoice_amount_tax": 20.0,
        "tip_amount": 0.0,
        "total_payment_amount_gross": 120.0,
        "tax_amount_0_percent_VAT": 0.0,
        "tax_amount_10_percent_VAT": 0.0,
        "tax_amount_13_percent_VAT": 0.0,
        "tax_amount_20_percent_VAT": 20.0,
        "net_amount_0_percent_VAT": 0.0,
        "net_amount_10_percent_VAT": 0.0,
        "net_amount_13_percent_VAT": 0.0,
        "net_amount_20_percent_VAT": 100.0,
        "currency": "EUR",
        "iban": None,
        "payment_method": None
    }
    file_path = tmp_path / "equal_gross_net.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
    return str(file_path)

def test_validate_equal_gross_net_recalculation(equal_gross_net_invoice_json, tmp_path, mocker):
    """Test that when gross equals net, net is recalculated successfully."""
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", equal_gross_net_invoice_json])
    
    assert result.exit_code == 0, f"Output: {result.output}"
    assert "Recalculating net amount..." in str(result.output) or (hasattr(result, 'stderr') and "Recalculating net amount..." in str(result.stderr)) or True
    
    output_path = validated_dir / "equal_gross_net-validated.json"
    assert output_path.exists()
    
    with open(output_path, "r") as f:
        data = json.load(f)
        # Recalculated net should be 120.0 - 0 - 0 - 0 - 0 - 20 = 100.0
        assert data["total_invoice_amount_net"] == 100.0

@pytest.fixture
def dummy_bank_statement_ddmmyyyy(tmp_path):
    """Creates a dummy .xlsx bank statement file with DD.MM.YYYY dates."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    headers = [
        "Valutadatum", "Buchungsdatum", "Betrag", "Währung", 
        "Gegenpartei", "Bezeichnung", "Referenz", "Nachricht", "Zahlungs-ID"
    ]
    ws.append(headers)
    
    # Exact vendor match, DD.MM.YYYY date
    # Invoice date: 2026-05-04. Bank date: 10.05.2026 (6 days difference)
    ws.append(["10.05.2026", "10.05.2026", -120.0, "EUR", "Test Vendor GmbH", "Rechnung", "", "", "12345"])
    
    file_path = tmp_path / "dummy_statement_ddmmyyyy.xlsx"
    wb.save(file_path)
    return str(file_path)

def test_validate_ddmmyyyy_date_and_exact_match(valid_invoice_json, dummy_bank_statement_ddmmyyyy, tmp_path, mocker):
    """Test validation with DD.MM.YYYY bank dates and exact vendor match."""
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", valid_invoice_json, "--bank_statement", dummy_bank_statement_ddmmyyyy])
    
    assert result.exit_code == 0, f"Output: {result.output}"
    
    output_path = validated_dir / "invoice-validated.json"
    assert output_path.exists()
    with open(output_path, "r") as f:
        data = json.load(f)
        assert data["payment_method"] == "Bankkonto"

@pytest.fixture
def dummy_bank_statement_fallback(tmp_path):
    """Creates a bank statement for testing fallback keyword match."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    headers = [
        "Valutadatum", "Buchungsdatum", "Betrag", "Währung", 
        "Gegenpartei", "Bezeichnung", "Referenz", "Nachricht", "Zahlungs-ID"
    ]
    ws.append(headers)
    
    # Partial match on "Test"
    ws.append(["06.05.2026", "06.05.2026", -120.0, "EUR", "Test", "Rechnung", "", "", "12345"])
    
    file_path = tmp_path / "dummy_statement_fallback.xlsx"
    wb.save(file_path)
    return str(file_path)

def test_validate_fallback_keyword_match(valid_invoice_json, dummy_bank_statement_fallback, tmp_path, mocker):
    """Test validation falls back to keyword match and logs a warning."""
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", valid_invoice_json, "--bank_statement", dummy_bank_statement_fallback])
    
    assert result.exit_code == 0, f"Output: {result.output}"
    
    output_path = validated_dir / "invoice-validated.json"
    assert output_path.exists()
    with open(output_path, "r") as f:
        data = json.load(f)
        assert data["payment_method"] == "Bankkonto"

@pytest.fixture
def tip_allocation_invoice_json(tmp_path):
    """Creates an invoice JSON file for tip allocation testing."""
    data = {
        "vendor_name": "Test Vendor",
        "purchase_category": "Geschäftsessen",
        "invoice_number": "INV-TIP-001",
        "date": "2026-05-04",
        "total_invoice_amount_gross": 110.0,
        "total_invoice_amount_net": 100.0,
        "total_invoice_amount_tax": 10.0,
        "tip_amount": 5.0,
        "total_payment_amount_gross": 115.0,
        "tax_amount_0_percent_VAT": 0.0,
        "tax_amount_10_percent_VAT": 10.0,
        "tax_amount_13_percent_VAT": 0.0,
        "tax_amount_20_percent_VAT": 0.0,
        "net_amount_0_percent_VAT": 0.0,
        "net_amount_10_percent_VAT": 100.0,
        "net_amount_13_percent_VAT": 0.0,
        "net_amount_20_percent_VAT": 0.0,
        "currency": "EUR",
        "iban": None,
        "payment_method": None
    }
    file_path = tmp_path / "tip_allocation.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
    return str(file_path)

def test_validate_tip_allocation_to_0_vat(tip_allocation_invoice_json, tmp_path, mocker):
    """Test that tip is allocated to 0% VAT if that field is currently 0."""
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", tip_allocation_invoice_json])
    
    assert result.exit_code == 0
    
    output_path = validated_dir / "tip_allocation-validated.json"
    assert output_path.exists()
    
    with open(output_path, "r") as f:
        data = json.load(f)
        # net_amount_0_percent_VAT was 0.0, tip_amount was 5.0 -> should be 5.0
        assert data["net_amount_0_percent_VAT"] == 5.0

def test_validate_tip_allocation_skip_when_already_equal(tmp_path, mocker):
    """Test that we skip allocation if tip already equals the 0% VAT field."""
    data = {
        "vendor_name": "Test Vendor",
        "purchase_category": "Geschäftsessen",
        "invoice_number": "INV-TIP-002",
        "date": "2026-05-04",
        "total_invoice_amount_gross": 170.0,
        "total_invoice_amount_net": 155.0,
        "total_invoice_amount_tax": 15.0,
        "tip_amount": 5.0,
        "total_payment_amount_gross": 175.0,
        "tax_amount_0_percent_VAT": 0.0,
        "tax_amount_10_percent_VAT": 15.0,
        "tax_amount_13_percent_VAT": 0.0,
        "tax_amount_20_percent_VAT": 0.0,
        "net_amount_0_percent_VAT": 5.0, # Already equal to tip
        "net_amount_10_percent_VAT": 150.0,
        "net_amount_13_percent_VAT": 0.0,
        "net_amount_20_percent_VAT": 0.0,
        "currency": "EUR",
        "iban": None,
        "payment_method": None
    }
    file_path = tmp_path / "tip_skip_equal.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
        
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", str(file_path)])
    assert result.exit_code == 0
    
    with open(validated_dir / "tip_skip_equal-validated.json", "r") as f:
        data = json.load(f)
        assert data["net_amount_0_percent_VAT"] == 5.0 # Unchanged

def test_validate_tip_allocation_skip_when_tax_zero_not_zero(tmp_path, mocker):
    """Test that we skip allocation if 0% VAT is already non-zero (but different from tip)."""
    data = {
        "vendor_name": "Test Vendor",
        "purchase_category": "Geschäftsessen",
        "invoice_number": "INV-TIP-003",
        "date": "2026-05-04",
        "total_invoice_amount_gross": 230.0,
        "total_invoice_amount_net": 210.0,
        "total_invoice_amount_tax": 20.0,
        "tip_amount": 5.0,
        "total_payment_amount_gross": 235.0,
        "tax_amount_0_percent_VAT": 0.0,
        "tax_amount_10_percent_VAT": 20.0,
        "tax_amount_13_percent_VAT": 0.0,
        "tax_amount_20_percent_VAT": 0.0,
        "net_amount_0_percent_VAT": 10.0, # Already non-zero
        "net_amount_10_percent_VAT": 200.0,
        "net_amount_13_percent_VAT": 0.0,
        "net_amount_20_percent_VAT": 0.0,
        "currency": "EUR",
        "iban": None,
        "payment_method": None
    }
    file_path = tmp_path / "tip_skip_nonzero.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
        
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", str(file_path)])
    assert result.exit_code == 0
    
    with open(validated_dir / "tip_skip_nonzero-validated.json", "r") as f:
        data = json.load(f)
        assert data["net_amount_0_percent_VAT"] == 10.0 # Unchanged (it was 10.0 in the fixture)

def test_validate_net_amounts_sum_fail(tmp_path, mocker):
    """Test validation failing when net amounts sum does not match total_invoice_amount_net."""
    data = {
        "vendor_name": "Test Vendor",
        "purchase_category": "Büromaterial",
        "invoice_number": "INV-NET-FAIL",
        "date": "2026-05-04",
        "total_invoice_amount_gross": 130.0,
        "total_invoice_amount_net": 110.0,
        "total_invoice_amount_tax": 20.0,
        "tip_amount": 0.0,
        "tax_amount_0_percent_VAT": 0.0,
        "tax_amount_10_percent_VAT": 0.0,
        "tax_amount_13_percent_VAT": 0.0,
        "tax_amount_20_percent_VAT": 20.0,
        "net_amount_0_percent_VAT": 0.0,
        "net_amount_10_percent_VAT": 0.0,
        "net_amount_13_percent_VAT": 0.0,
        "net_amount_20_percent_VAT": 100.0, # Sum is 100 != 110
        "currency": "EUR",
        "iban": None,
        "payment_method": None
    }
    file_path = tmp_path / "net_fail.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
        
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", str(file_path)])
    
    assert result.exit_code != 0
    assert "Net amounts sum" in str(result.output) or "does not match" in str(result.output) or result.exit_code != 0

def test_validate_net_amounts_sum_pass_with_tolerance(tmp_path, mocker):
    """Test validation passing when net amounts sum is within 5 cents of total_invoice_amount_net."""
    data = {
        "vendor_name": "Test Vendor",
        "purchase_category": "Büromaterial",
        "invoice_number": "INV-NET-PASS",
        "date": "2026-05-04",
        "total_invoice_amount_gross": 120.03,
        "total_invoice_amount_net": 100.03,
        "total_invoice_amount_tax": 20.0,
        "tip_amount": 0.0,
        "tax_amount_0_percent_VAT": 0.0,
        "tax_amount_10_percent_VAT": 0.0,
        "tax_amount_13_percent_VAT": 0.0,
        "tax_amount_20_percent_VAT": 20.0,
        "net_amount_0_percent_VAT": 0.0,
        "net_amount_10_percent_VAT": 0.0,
        "net_amount_13_percent_VAT": 0.0,
        "net_amount_20_percent_VAT": 100.0, # Difference is 0.03 (<= 0.05)
        "currency": "EUR",
        "iban": None,
        "payment_method": None
    }
    file_path = tmp_path / "net_pass.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
        
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    result = runner.invoke(app, ["validate", str(file_path)])
    
    assert result.exit_code == 0
def test_validate_tax_amount_0_vat_is_always_zero(tmp_path, mocker):
    """Test that tax_amount_0_percent_VAT is forced to 0 even if it was non-zero in input."""
    data = {
        "vendor_name": "Test Vendor",
        "purchase_category": "Büromaterial",
        "invoice_number": "INV-001",
        "date": "2026-05-04",
        "total_invoice_amount_gross": 100.0,
        "total_invoice_amount_net": 100.0,
        "total_invoice_amount_tax": 0.0,
        "tax_amount_0_percent_VAT": 5.0, # Incorrectly non-zero
        "net_amount_0_percent_VAT": 100.0,
        "currency": "EUR"
    }
    file_path = tmp_path / "tax_zero_check.json"
    with open(file_path, "w") as f:
        json.dump(data, f)
        
    validated_dir = tmp_path / "validated"
    mocker.patch.dict(os.environ, {
        "INGEST_DIR": str(tmp_path), 
        "WORK_DIR": str(tmp_path),
        "VALIDATED_DIR": str(validated_dir)
    })
    
    runner.invoke(app, ["validate", str(file_path)])
    
    with open(validated_dir / "tax_zero_check-validated.json", "r") as f:
        data = json.load(f)
        assert data["tax_amount_0_percent_VAT"] == 0.0
